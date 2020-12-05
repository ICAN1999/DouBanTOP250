# *_* coding: UTF-8 *_*
# 作者 : 吴圣冬
# 开发时间 : 2020/9/6  18:43
# 文件名称 : douban.PY
# 开发工具 : PyCharm
import requests
from bs4 import BeautifulSoup
import random
from openpyxl import Workbook


def get_html(url):
    # 使用代理
    iplist = ["121.232.148.225", "123.101.207.185", "69.195.157.162", "175.44.109.246", "175.42.128.246"]
    proxies = {"http": random.choice(iplist)}
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 \
        (KABUL, like Gecko) Chrome/85.0.4183.121 Safari/537.36 Edg/85.0.564.63',
        'cookie': 'bid=mDUxIx660Vg; douban-fav-remind=1; ll="118300"; '
                  '_vwo_uuid_v2=D0D095AA577CA9982D96BF53E5B82D902|19baada46789b48f67201d5ad830a0f6; '
                  '__yadk_uid=gAC153pUujwGBubBMhpUqg8osqjO7FfD; '
                  '__utmz=30149280.1601027331.15.8.utmcsr=accounts.douban.com|utmccn=('
                  'referral)|utmcmd=referral|utmcct=/passport/login; '
                  '__utmz=223695111.1601028074.12.5.utmcsr=accounts.douban.com|utmccn=('
                  'referral)|utmcmd=referral|utmcct=/passport/login; push_noty_num=0; push_doumail_num=0; ct=y; '
                  'ap_v=0,6.0; __utmc=30149280; __utmc=223695111; '
                  '__utma=30149280.846971982.1594690041.1601194295.1601199438.18; '
                  '__utma=223695111.2022617817.1595329698.1601194295.1601199438.15; __utmb=223695111.0.10.1601199438; '
                  '_pk_ses.100001.4cf6=*; _pk_ref.100001.4cf6=%5B%22%22%2C%22%22%2C1601199438%2C%22https%3A%2F'
                  '%2Faccounts.douban.com%2Fpassport%2Flogin%3Fredir%3Dhttps%253A%252F%252Fmovie.douban.com'
                  '%252Ftop250%22%5D; dbcl2="216449506:1anURmdR9nE"; ck=LxNQ; __utmt=1; __utmv=30149280.21644; '
                  'douban-profile-remind=1; '
                  '__gads=ID=c502f7a27a33facd-22f60957b9c300e3:T=1601201244:S=ALNI_MbN9wSYWbg2k5f3fWucwuXpORbGNw; '
                  '__utmb=30149280.20.10.1601199438; '
                  '_pk_id.100001.4cf6=d35078454df4f375.1595329083.15.1601201286.1601197192.'}
    res = requests.get(url, headers=headers, proxies=proxies, timeout=30)
    # res = requests.get(url, headers=headers)
    return res


def parser_html(res):
    soup = BeautifulSoup(res.text, 'html.parser')

    # 电影名
    movies = []
    targets = soup.find_all("div", class_="hd")
    for each in targets:
        try:
            movies.append(each.a.span.text)
        finally:
            continue

    # 评分
    ranks = []
    targets = soup.find_all("span", class_="rating_num")
    for each in targets:
        try:
            ranks.append(each.text)
        finally:
            continue

    # 导演
    director = []
    targets = soup.find_all("div", class_="bd")
    for each in targets:
        try:
            director.append(each.p.text.split('\n')[1].strip().split('\xa0\xa0\xa0')[0].lstrip('导演: '))
        finally:
            continue

    # 主演
    leading_actor = []
    targets = soup.find_all("div", class_="bd")
    for each in targets:
        try:
            leading_actor.append(each.p.text.split('\n')[1].strip().split('\xa0\xa0\xa0')[1].split('主演:')[1])
        finally:
            continue
    if movies[0] == "肖申克的救赎":
        leading_actor.insert(24, "无")
    if movies[0] == "大闹天宫":
        leading_actor.insert(6, "无")
        leading_actor.insert(8, "无")
    if movies[0] == "美国往事":
        leading_actor.insert(5, "无")
    if movies[0] == "阳光姐妹淘":
        leading_actor.insert(23, "无")
    if movies[0] == "七武士":
        leading_actor.insert(10, "无")
    if movies[0] == "模仿游戏":
        leading_actor.insert(4, "无")
        leading_actor.insert(9, "无")
    if movies[0] == "罗生门":
        leading_actor.insert(2, "无")

    # 评价人数
    people = []
    targets = soup.find_all("div", class_="star")
    for each in targets:
        try:
            people.append(str(each.contents[7].text).rstrip("人评价"))
        finally:
            continue

    # 电影描述，最后两页分别有两篇电影没有描述
    sentence = []
    targets = soup.find_all("span", class_="inq")
    for each in targets:
        try:
            sentence.append(str(each.text))
        finally:
            continue
    if sentence[22] == "天使保护事件始末。":
        sentence.insert(9, "无")
        sentence.insert(14, "无")
    if sentence[22] == "生病的E.T.皮肤的颜色就像柿子饼。":
        sentence.insert(6, "无")
        sentence.insert(22, "无")

    # 年份
    year = []
    targets = soup.find_all("div", class_="bd")
    for each in targets:
        try:
            year.append(each.p.text.split('\n')[2].strip().split('\xa0')[0].strip("(中国大陆)"))
        finally:
            continue
    if movies[0] == "大闹天宫":
        year[0] = "1961 / 1964 / 1978 / 2004"
    # 国家
    country = []
    targets = soup.find_all("div", class_="bd")
    for each in targets:
        try:
            country.append(each.p.text.split('\n')[2].strip().split('\xa0')[2])
        finally:
            continue
    # 类型
    kinds = []
    targets = soup.find_all("div", class_="bd")
    for each in targets:
        try:
            kinds.append(each.p.text.split('\n')[2].strip().split('\xa0')[4])
        finally:
            continue
    result = []
    length = len(movies)
    for i in range(length):
        result.append([movies[i], ranks[i], director[i], leading_actor[i],
                       people[i], sentence[i], year[i], country[i], kinds[i]])
    return result


# 找出一共有多少个页面
def find_pages(res):
    soup = BeautifulSoup(res.text, 'html.parser')
    depth = soup.find('span', class_='next').previous_sibling.previous_sibling.text

    return int(depth)


def save_to_excel(result):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "名称"
    ws["B1"] = "评分"
    ws["c1"] = "导演"
    ws['d1'] = "主演"
    ws["e1"] = "评价人数"
    ws["f1"] = "电影描述"
    ws["g1"] = "年份"
    ws["h1"] = "国家"
    ws["i1"] = "类型"
    for each in result:
        ws.append(each)
    wb.save("豆瓣电影TOP250.xlsx")


def main():
    host = "https://movie.douban.com/top250"
    res = get_html(host)
    depth = find_pages(res)
    result = []
    for i in range(depth):
        url = host + '/?start=' + str(25 * i)
        res = get_html(url)
        result.extend(parser_html(res))
    save_to_excel(result)


if __name__ == "__main__":
    main()
